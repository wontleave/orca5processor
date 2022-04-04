from itertools import groupby

import openpyxl
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from os import path
from orca5_utils import Orca5


def orca5_to_excel(orca5_objs, excel_filename="default.xlsx"):
    """

    :param orca5_objs:
    :type orca5_objs:
    :param excel_filename:
    :type excel_filename:
    :return:
    :rtype:
    """


def orca_to_excel(orca_objs, excel_filename="default.xlsx", temperature=298.15, write_nmr=False):
    """
    Convert the data in the list of Orca object to an excel spreadsheet

    Args:
    :param orca_objs: a list of Orca4 objects
    :type orca_objs: list
    :param temperature: the temperature which the thermochemical data is calculated
    :type temperature:
    Kwargs:
    :param excel_filename: the filename for the excel file
    :type excel_filename: str
    :param write_nmr: the flag to control whether to write NMR results to the excel file in a separate sheet
    :type write_nmr: bool

    """
    wb = Workbook()
    ws1 = wb.active
    to_be_written = []
    extra = 0
    row_counter = 1

    # sort the label
    labels = ["zpe_corr", "thermal_corr", "enthalpy_corr", "entropy_corr"]
    thermo_labels = ["zpe corrected energy", "total thermal energy", "total enthalpy", "final free energy"]
    all_labels = []
    # Create labels for all orca objects
    for _orca in orca_objs:
        temp = []
        if _orca.has_opt:
            temp = labels.copy()
            temp.append(_orca.methods[0].get_level_of_theory())
            for i in range(4):
                temp.append(f"{_orca.methods[0].get_level_of_theory()}+{thermo_labels[i]}")
            for sp in _orca.extra_single_points:
                temp.append(sp.get_level_of_theory())
                for i in range(4):
                    temp.append(f"{sp.get_level_of_theory()}-{thermo_labels[i]}")
        else:
            for sp in _orca.extra_single_points:
                temp.append(f"{sp.get_level_of_theory()}")
            temp = labels + temp
        all_labels.append(temp)

    labels_lengths = np.array([len(labels) for labels in all_labels])
    with_most_labels = np.argmax(labels_lengths)
    labels = all_labels[with_most_labels]
    labels.append("Num of neg freq(s)")

    # write the label
    for i, label in enumerate(labels):
        _ = ws1.cell(column=1, row=i + 2, value="{0}".format(label))

    for i, item in enumerate(orca_objs):
        print(f"Working on {item.name_of_struct}")
        _ = ws1.cell(column=i + 2, row=1, value="{0}".format(str(item.name_of_struct).strip("[]\'")))
        label_counter = 0
        current_labels = all_labels[i]

        if item.has_opt and item.has_freq:
            if item.geo_opt[0].thermochemistry.temp == temperature:
                if current_labels[label_counter] == labels[label_counter]:
                    to_be_written.append(item.geo_opt[0].thermochemistry.zero_pt_corr)
                    to_be_written.append(item.geo_opt[0].thermochemistry.thermal_corr)
                    to_be_written.append(item.geo_opt[0].thermochemistry.enthalpy_corr)
                    to_be_written.append(item.geo_opt[0].thermochemistry.entropy_corr)
                    label_counter += 4
                else:
                    label_counter += 4
                    for k in range(4):
                        to_be_written.append(0.0)
            else:
                key, value = next(iter(item.extra_thermo.items()))
                if key == temperature and value is not None and current_labels[label_counter] == labels[label_counter]:
                    to_be_written.append(item.extra_thermo[temperature].zero_pt_corr)
                    to_be_written.append(item.extra_thermo[temperature].thermal_corr)
                    to_be_written.append(item.extra_thermo[temperature].enthalpy_corr)
                    to_be_written.append(item.extra_thermo[temperature].entropy_corr)
                    label_counter += 4
                else:
                    label_counter += 4
                    for k in range(4):
                        to_be_written.append(0.0)
        else:
            for j in range(4):
                label_counter += 1
                to_be_written.append(0.0)

        if item.has_opt:
            if current_labels[label_counter] == labels[label_counter]:
                to_be_written.append(item.methods[0].hamiltonian.elec_energy)
                label_counter += 1
            else:
                to_be_written.append(0.0)
                label_counter += 1
        for sp in item.extra_single_points:
            if current_labels[label_counter] == labels[label_counter]:
                to_be_written.append(sp.hamiltonian.elec_energy)
                label_counter += 1
            else:
                to_be_written.append(0.0)
                label_counter += 1

        for j in range(len(to_be_written)):
            row_counter = j + 2 + extra * 4
            _ = ws1.cell(column=i + 2, row=row_counter, value=float("{0}".format(to_be_written[j])))
            if j >= 4:
                corr_energy = to_be_written[j]
                if item.has_opt:
                    for k in range(4):
                        row_counter = j + 2 + k + 1 + extra * 4
                        corr_energy += to_be_written[k]
                        _ = ws1.cell(column=i + 2, row=row_counter, value=float("{0}".format(corr_energy)))
                    extra += 1

        if item.has_opt and item.has_freq:
            _ = ws1.cell(column=i + 2, row=row_counter + 1,
                         value=int("{0}".format(item.geo_opt[0].frequencies.num_of_neg_freq)))
        extra = 0
        to_be_written = []
        ws1.column_dimensions[get_column_letter(i + 2)].width = 15

    if write_nmr:
        ws_iso = wb.create_sheet("NMRiso")
        ws_aniso = wb.create_sheet("NMRansio")
        all_labels = []

        for _orca in orca_objs:
            labels = []
            for _item in _orca.nmrs[0].nmr_atoms:
                labels.append(f"{_item[1]}{_item[0]}")
            all_labels.append(labels.copy())

        len_of_all_labels = [len(labels) for labels in all_labels]
        temp = list(groupby(len_of_all_labels))
        if len(temp) == 1:
            for idx in range(len_of_all_labels[0]):
                temp = [label[idx] for label in all_labels]
                if len(list(groupby(temp))) > 1:  # TODO where is groupby from pd or Itertool?
                    print(f"WARNING inconsistent labeling found in orca objects {labels} ... "
                          f"excel file will NOT be writtent")
                    break
        else:
            print(f"WARNING inconsistent labeling found in orca objects {temp} ... ")

        for i, label in enumerate(all_labels[0]):
            _ = ws_iso.cell(column=1, row=i + 2, value="{0}".format(label))
            _ = ws_aniso.cell(column=1, row=i + 2, value="{0}".format(label))

        for i, _orca in enumerate(orca_objs):
            _ = ws_iso.cell(column=i + 2, row=1, value="{0}".format(str(_orca.name_of_struct).strip("[]\'")))
            _ = ws_aniso.cell(column=i + 2, row=1, value="{0}".format(str(_orca.name_of_struct).strip("[]\'")))
            for j, nmr_tuple in enumerate(_orca.nmrs[0].nmr_atoms):
                _ = ws_iso.cell(column=i + 2, row=j + 2, value=float("{0}".format(nmr_tuple[2])))
                _ = ws_aniso.cell(column=i + 2, row=j + 2, value=float("{0}".format(nmr_tuple[2])))

    wb.save(excel_filename)
    wb.close()


def modify_orca_input(input_path, **kwargs):
    """

    :param input_path:
    :param kwargs: a dictionary of input keywords to modify
    :return:
    """
    with open(input_path, "r") as f:
        lines = f.readlines()

    change_max_iter = False
    if "recalc_hess" in kwargs:
        new_max_iter = kwargs["recalc_hess"] * 2 - 1
        change_max_iter = True

    new_input = ""

    for line in lines:
        if "recalc_hess" in line.lower():
            new_recalc_hess = kwargs["recalc_hess"]
            new_input += f"\trecalc_hess {new_recalc_hess}\n"
        elif change_max_iter and "maxiter" in line.lower():
            new_input += f"\tMaxIter {new_max_iter}\n"
            change_max_iter = False
        elif "xyzfile" in line.lower():
            if kwargs["xyz_path"] not in line:
                temp_ = line.strip().split()
                xyz_name = kwargs["xyz_path"]
                new_xyz_path = f"{path.dirname(temp_[-1])}/{xyz_name}"  # Only linux for now
                new_input += f"{temp_[0]} {temp_[1]} {temp_[2]} {temp_[3]} {new_xyz_path}\n"
        else:
            if "#" not in line and line != "\n" and line != "\t\n":
                if "!" in line or "%" in line or "*" in line:
                    new_input += f"{line.strip()}\n"
                elif "end" in line.lower():
                    temp_ = line.strip()
                    if temp_.lower() == "end":
                        new_input += f"{temp_}\n"
                else:
                    new_input += f"\t{line.strip()}\n"

    with open(input_path, "w") as f:
        f.writelines(new_input)